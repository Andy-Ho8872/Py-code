import openpyxl as xl

def process_workbook(filename):

    wb = xl.load_workbook(filename)
    sheet = wb['工作表1']

    for row in range(2, sheet.max_row + 1):
        cell = sheet.cell(row, column = 3) # row 若沒有指定數值的話，則印出所有row   ，第二個數字是 Column
        corrected_price = cell.value * 0.9
        corrected_price_cell = sheet.cell(row, 4)
        corrected_price_cell.value = corrected_price
        
    wb.save(filename)  # 儲存檔案(以另存新檔的方式) , 若沒有打檔名則會複寫原本的檔案
    print("Data Updated")

process_workbook('000.xlsx')